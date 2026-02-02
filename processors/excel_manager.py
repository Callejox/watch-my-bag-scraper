"""
Gestor de reportes Excel para exportar ventas detectadas.
Genera archivos mensuales con hojas separadas por plataforma.
"""

from datetime import date, datetime
from typing import List, Dict, Any, Optional
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from loguru import logger

import sys
sys.path.insert(0, str(Path(__file__).parent.parent))

from config import EXPORTS_DIR
from database.db_manager import DatabaseManager


class ExcelManager:
    """
    Genera y actualiza reportes Excel con las ventas detectadas.

    Estructura del archivo:
    - Hoja 1: "Inventario Chrono24" - Inventario de relojes Chrono24
    - Hoja 2: "Vendidos Chrono24" - Ventas de Chrono24
    - Hoja 3: "Inventario Vestiaire" - Inventario de Vestiaire
    - Hoja 4: "Vendidos Vestiaire" - Ventas de Vestiaire
    - Hoja 5: "Inventario Catawiki" - Inventario de subastas Catawiki
    - Hoja 6: "Vendidos Catawiki" - Ventas de Catawiki
    - Hoja 7: "Resumen Diario" - Estadísticas por día
    """

    # Estilos
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

    CURRENCY_FORMAT = '#,##0.00 "€"'
    DATE_FORMAT = "DD/MM/YYYY"

    BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    def __init__(self, db_manager: DatabaseManager = None):
        self.db = db_manager or DatabaseManager()
        self.exports_dir = EXPORTS_DIR
        self.exports_dir.mkdir(parents=True, exist_ok=True)
        self.logger = logger.bind(module="excel_manager")

    def _get_report_filename(self, year: int, month: int) -> Path:
        """
        Genera el nombre del archivo de reporte para un mes.

        Args:
            year: Año
            month: Mes

        Returns:
            Path del archivo
        """
        now = datetime.now()
        return self.exports_dir / f"ventas_detectadas_{year}-{month:02d}_{now.hour:02d}{now.minute:02d}.xlsx"

    def _apply_header_style(self, ws, row: int, num_cols: int):
        """Aplica estilos a la fila de encabezados."""
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.BORDER

    def _apply_data_style(self, ws, start_row: int, end_row: int, num_cols: int):
        """Aplica estilos a las filas de datos."""
        for row in range(start_row, end_row + 1):
            for col in range(1, num_cols + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = self.BORDER
                cell.alignment = Alignment(vertical="center")

    def _auto_adjust_columns(self, ws):
        """Ajusta automáticamente el ancho de las columnas."""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Max 50 caracteres
            ws.column_dimensions[column_letter].width = adjusted_width

    def create_chrono24_inventory_sheet(self, wb: Workbook, inventory: List[Dict[str, Any]]) -> None:
        """
        Crea la hoja de inventario actual de Chrono24.

        Args:
            wb: Workbook de Excel
            inventory: Lista de items en inventario
        """
        ws = wb.create_sheet("Inventario Chrono24", 0)

        # Encabezados
        headers = [
            "ID",
            "Modelo",
            "Referencia",
            "Precio (€)",
            "Fecha Publicación",
            "Condición",
            "Año Producción",
            "Material Caja",
            "Color Esfera",
            "País Vendedor",
            "URL Imagen",
            "URL"
        ]

        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        self._apply_header_style(ws, 1, len(headers))

        # Datos
        for row_idx, item in enumerate(inventory, 2):
            ws.cell(row=row_idx, column=1, value=item.get('listing_id', ''))
            ws.cell(row=row_idx, column=2, value=item.get('specific_model', ''))
            ws.cell(row=row_idx, column=3, value=item.get('reference_number', ''))

            price_cell = ws.cell(row=row_idx, column=4, value=item.get('listing_price'))
            price_cell.number_format = self.CURRENCY_FORMAT

            ws.cell(row=row_idx, column=5, value=item.get('upload_date', ''))
            ws.cell(row=row_idx, column=6, value=item.get('condition', ''))
            ws.cell(row=row_idx, column=7, value=item.get('year_of_production', ''))
            ws.cell(row=row_idx, column=8, value=item.get('case_material', ''))
            ws.cell(row=row_idx, column=9, value=item.get('dial_color', ''))
            ws.cell(row=row_idx, column=10, value=item.get('seller_location', ''))
            ws.cell(row=row_idx, column=11, value=item.get('image_url', ''))
            ws.cell(row=row_idx, column=12, value=item.get('url', ''))

        if inventory:
            self._apply_data_style(ws, 2, len(inventory) + 1, len(headers))

        self._auto_adjust_columns(ws)

    def create_chrono24_sales_sheet(self, wb: Workbook, sales: List[Dict[str, Any]]) -> None:
        """
        Crea la hoja de ventas de Chrono24.

        Args:
            wb: Workbook de Excel
            sales: Lista de ventas
        """
        ws = wb.create_sheet("Vendidos Chrono24", 1)

        # Encabezados
        headers = [
            "Fecha Detección",
            "Modelo",
            "Referencia",
            "Precio Venta (€)",
            "Fecha Publicación",
            "Días en Venta",
            "Condición",
            "Año Producción",
            "País Vendedor",
            "URL Imagen",
            "URL"
        ]

        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        self._apply_header_style(ws, 1, len(headers))

        # Datos
        for row_idx, sale in enumerate(sales, 2):
            ws.cell(row=row_idx, column=1, value=sale.get('detection_date', ''))
            ws.cell(row=row_idx, column=2, value=sale.get('specific_model', ''))
            ws.cell(row=row_idx, column=3, value=sale.get('reference_number', ''))

            price_cell = ws.cell(row=row_idx, column=4, value=sale.get('sale_price'))
            price_cell.number_format = self.CURRENCY_FORMAT

            ws.cell(row=row_idx, column=5, value=sale.get('upload_date', ''))
            ws.cell(row=row_idx, column=6, value=sale.get('days_on_sale', ''))
            ws.cell(row=row_idx, column=7, value=sale.get('condition', ''))
            ws.cell(row=row_idx, column=8, value=sale.get('year_of_production', ''))
            ws.cell(row=row_idx, column=9, value=sale.get('seller_location', ''))
            ws.cell(row=row_idx, column=10, value=sale.get('image_url', ''))
            ws.cell(row=row_idx, column=11, value=sale.get('url', ''))

        if sales:
            self._apply_data_style(ws, 2, len(sales) + 1, len(headers))

        self._auto_adjust_columns(ws)

    def create_vestiaire_inventory_sheet(self, wb: Workbook, inventory: List[Dict[str, Any]]) -> None:
        """
        Crea la hoja de inventario actual de Vestiaire.

        Args:
            wb: Workbook de Excel
            inventory: Lista de items en inventario
        """
        ws = wb.create_sheet("Inventario Vestiaire", 2)

        # Encabezados
        headers = [
            "ID",
            "Vendedor",
            "Descripción",
            "Precio (€)",
            "Fecha Publicación",
            "Condición",
            "URL Imagen",
            "URL"
        ]

        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        self._apply_header_style(ws, 1, len(headers))

        # Datos
        for row_idx, item in enumerate(inventory, 2):
            ws.cell(row=row_idx, column=1, value=item.get('listing_id', ''))
            ws.cell(row=row_idx, column=2, value=item.get('seller_id', ''))

            # Combinar marca + producto para descripción más completa
            brand = item.get('brand', '')
            product = item.get('product_name') or item.get('specific_model', '')
            if brand and product and brand not in product:
                description = f"{brand} - {product}"
            else:
                description = product or brand
            ws.cell(row=row_idx, column=3, value=description)

            price_cell = ws.cell(row=row_idx, column=4, value=item.get('listing_price'))
            price_cell.number_format = self.CURRENCY_FORMAT

            ws.cell(row=row_idx, column=5, value=item.get('upload_date', ''))
            ws.cell(row=row_idx, column=6, value=item.get('condition', ''))
            ws.cell(row=row_idx, column=7, value=item.get('image_url', ''))
            ws.cell(row=row_idx, column=8, value=item.get('url', ''))

        if inventory:
            self._apply_data_style(ws, 2, len(inventory) + 1, len(headers))

        self._auto_adjust_columns(ws)

    def create_vestiaire_sales_sheet(self, wb: Workbook, sales: List[Dict[str, Any]], index: int = 3) -> None:
        """
        Crea la hoja de ventas de Vestiaire.

        Args:
            wb: Workbook de Excel
            sales: Lista de ventas
            index: Posición de la hoja
        """
        ws = wb.create_sheet("Vendidos Vestiaire", index)

        # Encabezados
        headers = [
            "Fecha Detección",
            "Vendedor",
            "Descripción",
            "Precio Venta (€)",
            "Fecha Publicación",
            "Días en Venta",
            "Condición",
            "URL Imagen",
            "URL"
        ]

        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        self._apply_header_style(ws, 1, len(headers))

        # Datos
        for row_idx, sale in enumerate(sales, 2):
            ws.cell(row=row_idx, column=1, value=sale.get('detection_date', ''))
            ws.cell(row=row_idx, column=2, value=sale.get('seller_id', ''))

            # Combinar marca + producto para descripción
            brand = sale.get('brand', '')
            product = sale.get('product_name') or sale.get('specific_model', '')
            if brand and product and brand not in product:
                description = f"{brand} - {product}"
            else:
                description = product or brand
            ws.cell(row=row_idx, column=3, value=description)

            price_cell = ws.cell(row=row_idx, column=4, value=sale.get('sale_price'))
            price_cell.number_format = self.CURRENCY_FORMAT

            ws.cell(row=row_idx, column=5, value=sale.get('upload_date', ''))
            ws.cell(row=row_idx, column=6, value=sale.get('days_on_sale', ''))
            ws.cell(row=row_idx, column=7, value=sale.get('condition', ''))
            ws.cell(row=row_idx, column=8, value=sale.get('image_url', ''))
            ws.cell(row=row_idx, column=9, value=sale.get('url', ''))

        if sales:
            self._apply_data_style(ws, 2, len(sales) + 1, len(headers))

        self._auto_adjust_columns(ws)

    def create_catawiki_inventory_sheet(self, wb: Workbook, inventory: List[Dict[str, Any]], index: int = 4) -> None:
        """
        Crea la hoja de inventario actual de Catawiki.

        Args:
            wb: Workbook de Excel
            inventory: Lista de items en inventario
            index: Posición de la hoja
        """
        ws = wb.create_sheet("Inventario Catawiki", index)

        # Encabezados
        headers = [
            "ID",
            "Modelo",
            "Referencia",
            "Precio Actual (€)",
            "Fecha Publicación",
            "Condición",
            "Año Producción",
            "Material Caja",
            "Color Esfera",
            "País Vendedor",
            "URL Imagen",
            "URL"
        ]

        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        self._apply_header_style(ws, 1, len(headers))

        # Datos
        for row_idx, item in enumerate(inventory, 2):
            ws.cell(row=row_idx, column=1, value=item.get('listing_id', ''))
            ws.cell(row=row_idx, column=2, value=item.get('specific_model', ''))
            ws.cell(row=row_idx, column=3, value=item.get('reference_number', ''))

            price_cell = ws.cell(row=row_idx, column=4, value=item.get('listing_price'))
            price_cell.number_format = self.CURRENCY_FORMAT

            ws.cell(row=row_idx, column=5, value=item.get('upload_date', ''))
            ws.cell(row=row_idx, column=6, value=item.get('condition', ''))
            ws.cell(row=row_idx, column=7, value=item.get('year_of_production', ''))
            ws.cell(row=row_idx, column=8, value=item.get('case_material', ''))
            ws.cell(row=row_idx, column=9, value=item.get('dial_color', ''))
            ws.cell(row=row_idx, column=10, value=item.get('seller_location', ''))
            ws.cell(row=row_idx, column=11, value=item.get('image_url', ''))
            ws.cell(row=row_idx, column=12, value=item.get('url', ''))

        if inventory:
            self._apply_data_style(ws, 2, len(inventory) + 1, len(headers))

        self._auto_adjust_columns(ws)

    def create_catawiki_sales_sheet(self, wb: Workbook, sales: List[Dict[str, Any]], index: int = 5) -> None:
        """
        Crea la hoja de ventas de Catawiki.

        Args:
            wb: Workbook de Excel
            sales: Lista de ventas
            index: Posición de la hoja
        """
        ws = wb.create_sheet("Vendidos Catawiki", index)

        # Encabezados
        headers = [
            "Fecha Detección",
            "Modelo",
            "Referencia",
            "Precio Venta (€)",
            "Fecha Publicación",
            "Días en Venta",
            "Condición",
            "Año Producción",
            "País Vendedor",
            "URL Imagen",
            "URL"
        ]

        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        self._apply_header_style(ws, 1, len(headers))

        # Datos
        for row_idx, sale in enumerate(sales, 2):
            ws.cell(row=row_idx, column=1, value=sale.get('detection_date', ''))
            ws.cell(row=row_idx, column=2, value=sale.get('specific_model', ''))
            ws.cell(row=row_idx, column=3, value=sale.get('reference_number', ''))

            price_cell = ws.cell(row=row_idx, column=4, value=sale.get('sale_price'))
            price_cell.number_format = self.CURRENCY_FORMAT

            ws.cell(row=row_idx, column=5, value=sale.get('upload_date', ''))
            ws.cell(row=row_idx, column=6, value=sale.get('days_on_sale', ''))
            ws.cell(row=row_idx, column=7, value=sale.get('condition', ''))
            ws.cell(row=row_idx, column=8, value=sale.get('year_of_production', ''))
            ws.cell(row=row_idx, column=9, value=sale.get('seller_location', ''))
            ws.cell(row=row_idx, column=10, value=sale.get('image_url', ''))
            ws.cell(row=row_idx, column=11, value=sale.get('url', ''))

        if sales:
            self._apply_data_style(ws, 2, len(sales) + 1, len(headers))

        self._auto_adjust_columns(ws)

    def create_summary_sheet(
        self,
        wb: Workbook,
        chrono24_sales: List[Dict[str, Any]],
        vestiaire_sales: List[Dict[str, Any]],
        catawiki_sales: List[Dict[str, Any]] = None,
        index: int = 6
    ) -> None:
        """
        Crea la hoja de resumen diario.

        Args:
            wb: Workbook de Excel
            chrono24_sales: Ventas de Chrono24
            vestiaire_sales: Ventas de Vestiaire
            catawiki_sales: Ventas de Catawiki
            index: Posición de la hoja
        """
        catawiki_sales = catawiki_sales or []
        ws = wb.create_sheet("Resumen Diario", index)

        # Encabezados
        headers = [
            "Fecha",
            "Chrono24 Ventas",
            "Chrono24 Vol. (€)",
            "Vestiaire Ventas",
            "Vestiaire Vol. (€)",
            "Catawiki Ventas",
            "Catawiki Vol. (€)",
            "Total Ventas",
            "Total Vol. (€)"
        ]

        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        self._apply_header_style(ws, 1, len(headers))

        # Agrupar ventas por fecha
        daily_data = {}

        for sale in chrono24_sales:
            date_key = sale.get('detection_date', '')
            if date_key not in daily_data:
                daily_data[date_key] = {
                    'chrono24_count': 0, 'chrono24_volume': 0,
                    'vestiaire_count': 0, 'vestiaire_volume': 0,
                    'catawiki_count': 0, 'catawiki_volume': 0
                }
            daily_data[date_key]['chrono24_count'] += 1
            daily_data[date_key]['chrono24_volume'] += sale.get('sale_price', 0) or 0

        for sale in vestiaire_sales:
            date_key = sale.get('detection_date', '')
            if date_key not in daily_data:
                daily_data[date_key] = {
                    'chrono24_count': 0, 'chrono24_volume': 0,
                    'vestiaire_count': 0, 'vestiaire_volume': 0,
                    'catawiki_count': 0, 'catawiki_volume': 0
                }
            daily_data[date_key]['vestiaire_count'] += 1
            daily_data[date_key]['vestiaire_volume'] += sale.get('sale_price', 0) or 0

        for sale in catawiki_sales:
            date_key = sale.get('detection_date', '')
            if date_key not in daily_data:
                daily_data[date_key] = {
                    'chrono24_count': 0, 'chrono24_volume': 0,
                    'vestiaire_count': 0, 'vestiaire_volume': 0,
                    'catawiki_count': 0, 'catawiki_volume': 0
                }
            daily_data[date_key]['catawiki_count'] += 1
            daily_data[date_key]['catawiki_volume'] += sale.get('sale_price', 0) or 0

        # Ordenar por fecha
        sorted_dates = sorted(daily_data.keys())

        # Escribir datos
        for row_idx, date_key in enumerate(sorted_dates, 2):
            data = daily_data[date_key]
            total_count = data['chrono24_count'] + data['vestiaire_count'] + data['catawiki_count']
            total_volume = data['chrono24_volume'] + data['vestiaire_volume'] + data['catawiki_volume']

            ws.cell(row=row_idx, column=1, value=date_key)
            ws.cell(row=row_idx, column=2, value=data['chrono24_count'])

            vol_cell = ws.cell(row=row_idx, column=3, value=data['chrono24_volume'])
            vol_cell.number_format = self.CURRENCY_FORMAT

            ws.cell(row=row_idx, column=4, value=data['vestiaire_count'])

            vol_cell = ws.cell(row=row_idx, column=5, value=data['vestiaire_volume'])
            vol_cell.number_format = self.CURRENCY_FORMAT

            ws.cell(row=row_idx, column=6, value=data['catawiki_count'])

            vol_cell = ws.cell(row=row_idx, column=7, value=data['catawiki_volume'])
            vol_cell.number_format = self.CURRENCY_FORMAT

            ws.cell(row=row_idx, column=8, value=total_count)

            vol_cell = ws.cell(row=row_idx, column=9, value=total_volume)
            vol_cell.number_format = self.CURRENCY_FORMAT

        if daily_data:
            self._apply_data_style(ws, 2, len(daily_data) + 1, len(headers))

        # Fila de totales
        if daily_data:
            total_row = len(daily_data) + 2

            ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)

            total_c24_count = sum(d['chrono24_count'] for d in daily_data.values())
            total_c24_vol = sum(d['chrono24_volume'] for d in daily_data.values())
            total_vc_count = sum(d['vestiaire_count'] for d in daily_data.values())
            total_vc_vol = sum(d['vestiaire_volume'] for d in daily_data.values())
            total_cw_count = sum(d['catawiki_count'] for d in daily_data.values())
            total_cw_vol = sum(d['catawiki_volume'] for d in daily_data.values())

            ws.cell(row=total_row, column=2, value=total_c24_count).font = Font(bold=True)

            vol_cell = ws.cell(row=total_row, column=3, value=total_c24_vol)
            vol_cell.number_format = self.CURRENCY_FORMAT
            vol_cell.font = Font(bold=True)

            ws.cell(row=total_row, column=4, value=total_vc_count).font = Font(bold=True)

            vol_cell = ws.cell(row=total_row, column=5, value=total_vc_vol)
            vol_cell.number_format = self.CURRENCY_FORMAT
            vol_cell.font = Font(bold=True)

            ws.cell(row=total_row, column=6, value=total_cw_count).font = Font(bold=True)

            vol_cell = ws.cell(row=total_row, column=7, value=total_cw_vol)
            vol_cell.number_format = self.CURRENCY_FORMAT
            vol_cell.font = Font(bold=True)

            ws.cell(row=total_row, column=8, value=total_c24_count + total_vc_count + total_cw_count).font = Font(bold=True)

            vol_cell = ws.cell(row=total_row, column=9, value=total_c24_vol + total_vc_vol + total_cw_vol)
            vol_cell.number_format = self.CURRENCY_FORMAT
            vol_cell.font = Font(bold=True)

        self._auto_adjust_columns(ws)

    def generate_monthly_report(
        self,
        year: int = None,
        month: int = None
    ) -> Path:
        """
        Genera el reporte mensual completo.

        Hojas generadas:
        1. Inventario Chrono24 - Todos los items actuales
        2. Vendidos Chrono24 - Items que desaparecieron (vendidos)
        3. Inventario Vestiaire - Todos los items actuales
        4. Vendidos Vestiaire - Items que desaparecieron (vendidos)
        5. Inventario Catawiki - Todos los items actuales
        6. Vendidos Catawiki - Items que desaparecieron (vendidos)
        7. Resumen Diario - Estadísticas por día

        Args:
            year: Año (default: año actual)
            month: Mes (default: mes actual)

        Returns:
            Path del archivo generado
        """
        today = date.today()
        year = year or today.year
        month = month or today.month

        self.logger.info(f"Generando reporte para {year}-{month:02d}")

        # Obtener inventario actual
        chrono24_inventory = self.db.get_inventory_for_month(year, month, "chrono24")
        vestiaire_inventory = self.db.get_inventory_for_month(year, month, "vestiaire")
        catawiki_inventory = self.db.get_inventory_for_month(year, month, "catawiki")

        # Obtener ventas del mes
        chrono24_sales = self.db.get_sales_for_month(year, month, "chrono24")
        vestiaire_sales = self.db.get_sales_for_month(year, month, "vestiaire")
        catawiki_sales = self.db.get_sales_for_month(year, month, "catawiki")

        self.logger.info(
            f"Inventario: Chrono24={len(chrono24_inventory)}, "
            f"Vestiaire={len(vestiaire_inventory)}, "
            f"Catawiki={len(catawiki_inventory)}"
        )
        self.logger.info(
            f"Ventas: Chrono24={len(chrono24_sales)}, "
            f"Vestiaire={len(vestiaire_sales)}, "
            f"Catawiki={len(catawiki_sales)}"
        )

        # Crear workbook
        wb = Workbook()
        # Eliminar hoja por defecto
        wb.remove(wb.active)

        # Crear hojas (en orden)
        self.create_chrono24_inventory_sheet(wb, chrono24_inventory)
        self.create_chrono24_sales_sheet(wb, chrono24_sales)
        self.create_vestiaire_inventory_sheet(wb, vestiaire_inventory)
        self.create_vestiaire_sales_sheet(wb, vestiaire_sales)
        self.create_catawiki_inventory_sheet(wb, catawiki_inventory)
        self.create_catawiki_sales_sheet(wb, catawiki_sales)
        self.create_summary_sheet(wb, chrono24_sales, vestiaire_sales, catawiki_sales)

        # Guardar archivo
        filepath = self._get_report_filename(year, month)
        wb.save(filepath)

        self.logger.info(f"Reporte guardado en: {filepath}")
        return filepath

    def update_daily_report(self) -> Path:
        """
        Actualiza el reporte del mes actual con los datos de hoy.

        Returns:
            Path del archivo actualizado
        """
        today = date.today()
        return self.generate_monthly_report(today.year, today.month)

    def export_to_csv(
        self,
        year: int = None,
        month: int = None,
        platform: str = None
    ) -> List[Path]:
        """
        Exporta los datos a archivos CSV.

        Args:
            year: Año (default: año actual)
            month: Mes (default: mes actual)
            platform: Filtrar por plataforma (opcional)

        Returns:
            Lista de paths de archivos generados
        """
        today = date.today()
        year = year or today.year
        month = month or today.month

        files = []

        # Exportar Chrono24
        if platform is None or platform == "chrono24":
            chrono24_sales = self.db.get_sales_for_month(year, month, "chrono24")
            if chrono24_sales:
                df = pd.DataFrame(chrono24_sales)
                filepath = self.exports_dir / f"chrono24_ventas_{year}-{month:02d}.csv"
                df.to_csv(filepath, index=False, encoding='utf-8')
                files.append(filepath)
                self.logger.info(f"CSV Chrono24 guardado: {filepath}")

        # Exportar Vestiaire
        if platform is None or platform == "vestiaire":
            vestiaire_sales = self.db.get_sales_for_month(year, month, "vestiaire")
            if vestiaire_sales:
                df = pd.DataFrame(vestiaire_sales)
                filepath = self.exports_dir / f"vestiaire_ventas_{year}-{month:02d}.csv"
                df.to_csv(filepath, index=False, encoding='utf-8')
                files.append(filepath)
                self.logger.info(f"CSV Vestiaire guardado: {filepath}")

        # Exportar Catawiki
        if platform is None or platform == "catawiki":
            catawiki_sales = self.db.get_sales_for_month(year, month, "catawiki")
            if catawiki_sales:
                df = pd.DataFrame(catawiki_sales)
                filepath = self.exports_dir / f"catawiki_ventas_{year}-{month:02d}.csv"
                df.to_csv(filepath, index=False, encoding='utf-8')
                files.append(filepath)
                self.logger.info(f"CSV Catawiki guardado: {filepath}")

        return files

    def get_export_stats(self) -> Dict[str, Any]:
        """
        Obtiene estadísticas de los archivos exportados.

        Returns:
            Diccionario con información de archivos
        """
        files = list(self.exports_dir.glob("*.xlsx")) + list(self.exports_dir.glob("*.csv"))

        return {
            'export_directory': str(self.exports_dir),
            'total_files': len(files),
            'files': [
                {
                    'name': f.name,
                    'size_kb': f.stat().st_size / 1024,
                    'modified': datetime.fromtimestamp(f.stat().st_mtime).isoformat()
                }
                for f in sorted(files, key=lambda x: x.stat().st_mtime, reverse=True)
            ]
        }
